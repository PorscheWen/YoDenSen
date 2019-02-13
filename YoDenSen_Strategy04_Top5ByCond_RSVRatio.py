# -*- coding: utf-8 -*-
'''
Created on Mon Dec 31 14:35:01 2018

@author: User
'''

import numpy as np
import pandas as pd
from finlab.data import Data
import datetime
import os
from openpyxl import load_workbook

# Remove 'Fututre Warning'
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)


#def backtest(start_date, end_date, hold_days, strategy, data, weight='average', benchmark=None, stop_loss=None, stop_profit=None):
def backtest(start_date, end_date, hold_days, RSV_Ndays, RSV_Ratio, lstStockSorted, PickStockNumber, data, PathExcelFile, weight='average', benchmark=None, stop_loss=None, stop_profit=None):  
    # portfolio check
    if weight != 'average' and weight != 'price':
        print('Backtest stop, weight should be "average" or "price", find', weight, 'instead')

    # get price data in order backtest
    data.date = end_date
    price = data.get('收盤價', (end_date - start_date).days)
    # start from 1 TWD at start_date, 
    end = 1
    #date = start_date
    
    # record some history
    equality = pd.Series()
    nstock = {}
    transections = pd.DataFrame()
    maxreturn = -10000
    minreturn = 10000
    
    def trading_day(date):
        if date not in price.index:
            temp = price.loc[date:]
            if temp.empty:
                return price.index[-1]
            else:
                return temp.index[0]
        else:
            return date
    
    def date_iter_periodicity(start_date, end_date, hold_days):
        date = start_date
        while date < end_date:
            yield (date), (date + datetime.timedelta(hold_days))
            date += datetime.timedelta(hold_days)
                
    def date_iter_specify_dates(start_date, end_date, hold_days):
        dlist = [start_date] + hold_days + [end_date]
        if dlist[0] == dlist[1]:
            dlist = dlist[1:]
        if dlist[-1] == dlist[-2]:
            dlist = dlist[:-1]
        for sdate, edate in zip(dlist, dlist[1:]):
            yield (sdate), (edate)
    
    if isinstance(hold_days, int):
        dates = date_iter_periodicity(start_date, end_date, hold_days)
    elif isinstance(hold_days, list):
        dates = date_iter_specify_dates(start_date, end_date, hold_days)
    else:
        print('the type of hold_dates should be list or int.')
        return None

    for sdate, edate in dates:
        
        # select stocks at date
        data.date = sdate
        
        #2019.01.04 Bao Take strategy into stock list
        #stocks = YoDenSen_Adv_Sorted(data, RSV_Ndays)
        市值, 自由現金流, 股東權益報酬率, 營業利益成長率, 市值營收比, rsv = pdData(data, RSV_Ndays)
        # Choose conditioned and sorted stock
        dfSortedStock = YoDenSen_Adv_Sorted(lstStockSorted, 市值, 自由現金流, 股東權益報酬率, 營業利益成長率, 市值營收比, rsv, RSV_Ratio) 
        # Choose top 5 Stock ID
        dfSortedHeadStock = dfSortedStock.head(PickStockNumber)
        print('-'*10 +' Buy Stock ' + '-'*10)
        print(dfSortedHeadStock)
        
        # hold the stocks for hold_days day
        s = price[dfSortedHeadStock.index & price.columns][sdate:edate].iloc[1:]
        
        
        if s.empty:
            s = pd.Series(1, index=pd.date_range(sdate + datetime.timedelta(days=1), edate))
        else:
            
            if stop_loss != None:
                below_stop = ((s / s.bfill().iloc[0]) - 1)*100 < -np.abs(stop_loss)
                below_stop = (below_stop.cumsum() > 0).shift(2).fillna(False)
                s[below_stop] = np.nan
                
            if stop_profit != None:
                above_stop = ((s / s.bfill().iloc[0]) - 1)*100 > np.abs(stop_profit)
                above_stop = (above_stop.cumsum() > 0).shift(2).fillna(False)
                s[above_stop] = np.nan
                
            s.dropna(axis=1, how='all', inplace=True)
            
            # record transections
            transections = transections.append(pd.DataFrame({
                'buy_price': s.bfill().iloc[0],
                'sell_price': s.apply(lambda s:s.dropna().iloc[-1]),
                'lowest_price': s.min(),
                'highest_price': s.max(),
                'buy_date': pd.Series(s.index[0], index=s.columns),
                'sell_date': s.apply(lambda s:s.dropna().index[-1]),
            }))
            
            transections['profit(%)'] = (transections['sell_price'] / transections['buy_price'] - 1) * 100
            
            s.ffill(inplace=True)
                
            # calculate equality
            # normalize and average the price of each stocks
            if weight == 'average':
                s = s/s.bfill().iloc[0]
            s = s.mean(axis=1)
            s = s / s.bfill()[0]
        
        # print some log    
        #print(sdate,'-', edate, '報酬率: %.2f'%( s.iloc[-1]/s.iloc[0] * 100 - 100), '%', 'nstock', len(dfSortedHeadStock))
        strDate = str(sdate) +'to '+ str(edate)
        strDate = strDate.replace('00:00:00', '')
        strROE = '報酬率: %.2f'%( s.iloc[-1]/s.iloc[0] * 100 - 100)
        print(strDate+'\n'+strROE)
        
        # Write to excel        
        dfDateROE = pd.DataFrame([strDate, strROE], ['Date', 'ROE'])
        dfDateROE.columns = ['Date & ROE'] 
        dfDateROEStock = dfDateROE.append(dfSortedHeadStock)
        sheet_name = 'RSV'+str(RSV_Ndays)+'_RSVRatio'+str(RSV_Ratio)+'_Hold'+str(hold_days)
        AddDataToExcelSheet(PathExcelFile, sheet_name, dfDateROEStock)
        
        maxreturn = max(maxreturn, s.iloc[-1]/s.iloc[0] * 100 - 100)
        minreturn = min(minreturn, s.iloc[-1]/s.iloc[0] * 100 - 100)
        
        # plot backtest result
        ((s*end-1)*100).plot()
        equality = equality.append(s*end)
        end = (s/s[0]*end).iloc[-1]
        # Write ROE to Excel
        sheet_name = 'ALL_ROE'
        AddDataToExcelSheet(PathExcelFile, sheet_name, dfSumROE_All_Sorted)
        dicData = dic_ROE(equality, RSV_Ndays, RSV_Ratio, StockHold_Days)
        
        # add nstock history
        nstock[sdate] = len(dfSortedHeadStock)
        
    #print('每次換手最大報酬 : %.2f ％' % maxreturn)
    #print('每次換手最少報酬 : %.2f ％' % minreturn)
    
    
    if benchmark is None:
        benchmark = price['0050'][start_date:end_date].iloc[1:]
    
    '''
    # bechmark (thanks to Markk1227)
    ((benchmark/benchmark[0]-1)*100).plot(color=(0.8,0.8,0.8))
    plt.ylabel('Return On Investment (%)')
    plt.grid(linestyle='-.')
    plt.show()
    ((benchmark/benchmark.cummax()-1)*100).plot(legend=True, color=(0.8,0.8,0.8))
    ((equality/equality.cummax()-1)*100).plot(legend=True)
    plt.ylabel('Dropdown (%)')
    plt.grid(linestyle='-.')
    plt.show()
    pd.Series(nstock).plot.bar()
    plt.ylabel('Number of stocks held')
    '''    
    return equality, transections, dicData

def toSeasonal(df):
    season4 = df[df.index.month == 3]
    season1 = df[df.index.month == 5]
    season2 = df[df.index.month == 8]
    season3 = df[df.index.month == 11]

    season1.index = season1.index.year
    season2.index = season2.index.year
    season3.index = season3.index.year
    season4.index = season4.index.year - 1

    newseason1 = season1
    newseason2 = season2 - season1.reindex_like(season2)
    newseason3 = season3 - season2.reindex_like(season3)
    newseason4 = season4 - season3.reindex_like(season4)

    newseason1.index = pd.to_datetime(newseason1.index.astype(str) + '-05-15')
    newseason2.index = pd.to_datetime(newseason2.index.astype(str) + '-08-14')
    newseason3.index = pd.to_datetime(newseason3.index.astype(str) + '-11-14')
    newseason4.index = pd.to_datetime((newseason4.index + 1).astype(str) + '-03-31')

    return newseason1.append(newseason2).append(newseason3).append(newseason4).sort_index()


def pdData(data, RSV_Ndays):
    
    股本 = data.get('股本合計', 1)#.drop_duplicates(['stock_id', 'date'], keep='last')#.pivot(index='date', columns='stock_id')
    price = data.get('收盤價', 200) #先取出200筆
    當天股價 = price[:股本.index[-1]].iloc[-1] 
    當天股本 = 股本.iloc[-1]
    市值 = 當天股本 * 當天股價 / 10 * 1000
   

    df1 = toSeasonal(data.get('投資活動之淨現金流入（流出）', 5))
    df2 = toSeasonal(data.get('營業活動之淨現金流入（流出）', 5))
    自由現金流 = (df1 + df2).iloc[-4:].mean()
    
    
    稅後淨利 = data.get('本期淨利（淨損）', 1)
    
    # 股東權益，有兩個名稱，有些公司叫做權益總計，有些叫做權益總額
    # 所以得把它們抓出來
    權益總計 = data.get('權益總計', 1)
    權益總額 = data.get('權益總額', 1)
    
    # 並且把它們合併起來
    權益總計.fillna(權益總額, inplace=True)
        
    股東權益報酬率 = 稅後淨利.iloc[-1] / 權益總計.iloc[-1]
    
    
    營業利益 = data.get('營業利益（損失）', 5)
    營業利益成長率 = (營業利益.iloc[-1] / 營業利益.iloc[-5] - 1) * 100
    
    
    當月營收 = data.get('當月營收', 4) * 1000
    當季營收 = 當月營收.iloc[-4:].sum()
    市值營收比 = 市值 / 當季營收        
    
    # N天內最高
    rsv = (price.iloc[-1] - price.iloc[-RSV_Ndays:].min()) / (price.iloc[-RSV_Ndays:].max() - price.iloc[-RSV_Ndays:].min())
    
    return 市值, 自由現金流, 股東權益報酬率, 營業利益成長率, 市值營收比, rsv  

    
def YoDenSen_Adv_Sorted(lstStockSorted, 市值, 自由現金流, 股東權益報酬率, 營業利益成長率, 市值營收比, rsv, RSV_Ratio):  
    condition1 = (市值 < 1e10)
    condition2 = 自由現金流 > 0
    condition3 = 股東權益報酬率 > 0
    condition4 = 營業利益成長率 > 0.1
    condition5 = 市值營收比 < 3
    condition6 = rsv > RSV_Ratio
    
    # Select stocks according to condition 1 - 6
    select_stock = condition1 & condition2 & condition3 & condition4 & condition5 & condition6
    #print(select_stock)    
    #print(select_stock[select_stock])
    #input('select_stock')
    
    
    
    # Combine all stock information into one dataframe
    dfStockInfo = StockInfo(select_stock[select_stock], 市值, 自由現金流, 股東權益報酬率, 營業利益成長率, 市值營收比, rsv)    
    dfSortedStock = StockSorted(dfStockInfo, lstStockSorted)     
    #print(dfSortedStock)
    #input('dfSortedStock')
    
    # List all index with Conditions and Sorted
    #lstFilteredSortedStocks = dfSortedStock.index.values
    return dfSortedStock

    # Filtered Nan rows and sorted by lstStockSorted
def StockSorted(dfStockInfo, lstStockSorted, ):
    # Drop value is Nan by row
    dfFilteredStock = dfStockInfo.dropna()
    #print(dfFilteredStock)
    #input('dfFilteredStock')
    # Sort value by Column name
    dfSortedStock = dfFilteredStock.sort_values(by = lstStockSorted, ascending = False)
    return dfSortedStock
   

def StockInfo(符合資格, 市值, 自由現金流, 股東權益報酬率, 營業利益成長率, 市值營收比, rsv):
    dfStockInfo = pd.concat([符合資格, 市值, 自由現金流, 股東權益報酬率, 營業利益成長率, 市值營收比, rsv], axis = 1, sort = False)
    dfStockInfo.columns = ['符合條件', '市值', '自由現金流', '股東權益報酬率', '營業利益成長率', '市值營收比', 'rsv']
    #print(dfStockInfo)
    #input('dfStockInfo')
    return dfStockInfo
    

def PickDate(strYMD):
    # Input strYMD = '2018_01_31'
    dateYMDT = datetime.datetime.strptime(strYMD, '%Y_%m_%d').replace(tzinfo = None)
    dateYMD = dateYMDT.replace(tzinfo = None)
    return dateYMD


def dic_ROE(equality, RSV_Ndays, RSV_Ratio, StockHold_Days):
    SumROEPercent = (equality[-1]-1)*100
    SumROEPercent_Round = round(SumROEPercent, 2)
    print('總報酬率 : '+str(SumROEPercent_Round)+'%')    
    #dicData = {'RSV_Ndays': RSV_Ndays, 'StockHold_Days': StockHold_Days, 'SumROE': SumROEPercent_Round}
    dicData = {'RSV_Ndays': RSV_Ndays, 'RSV_Ratio': RSV_Ratio, 'StockHold_Days': StockHold_Days, 'SumROE': SumROEPercent_Round}
    #SumROE_DF = pd.DataFrame(dicData)
    return dicData

def AddDataToExcelSheet(PathExcelFile, sheet_name, dfData): 
    writer = pd.ExcelWriter(PathExcelFile, engine='openpyxl')     
    startrow = None
    truncate_sheet = ""
    try:
        # try to open an existing workbook
        writer.book = load_workbook(PathExcelFile)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row +2

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
    if startrow is None:
        startrow = 0
    # write out the new sheet
    dfData.to_excel(writer, sheet_name, startrow=startrow)
    # save the workbook
    writer.save()
    




if __name__ == '__main__':  
    
    # Sorted Condition by evaluated test from Strategy03
    lstStockSorted = ['市值', '自由現金流', 'rsv', '股東權益報酬率', '營業利益成長率', '市值營收比']
    
    # 最大股票數
    PickStockNumber = 5
    
    # 統計所有ROE的資料 DataFrame format
    SumROE_DF_All = pd.DataFrame()
    
    # 輸出Excel檔案位置 檔案名稱: 20190131_1359 (YYYYMMDD_HHMM)
    PathExcelFileName = datetime.datetime.now().strftime('%Y%m%d_%H%M')+'.xlsx'
    PathExcelFile = os.getcwd()+'\\'+PathExcelFileName
    
    for RSV_Ndays in range (50, 251, 50):
        for RSV_Ratio in range(0, 101, 10):
            RSV_Ratio = RSV_Ratio/100
            for StockHold_Days in range (15, 121, 15):
                # Inpu
                StartDate = PickDate('2016_01_01')
                EndDate = PickDate('2018_12_30')
                print('-'*10 + ' RSV & Hold Days '  +'-'*10)
                print('RSV_Ndays: ' + str(RSV_Ndays))
                print('RSV_Ratio: ' + str(RSV_Ratio))
                print('StockHold_Days: ' + str(StockHold_Days))
                
                # Evaluation:
                # equality是累計總報酬率  2018-11-30    0.859861
                # transections是買賣個股的價格 1701     2018-05-03      20.90     ...     2018-05-31       20.55
                equality, transections, dicData = backtest(StartDate, EndDate, StockHold_Days, RSV_Ndays, RSV_Ratio, lstStockSorted, PickStockNumber, Data(), PathExcelFile)
               
                
                # Output:
                SumROE_DF_All = SumROE_DF_All.append(dicData, ignore_index = True)
    dfSumROE_All_Sorted = SumROE_DF_All.sort_values( by = 'SumROE',  ascending= True)
    print('-'*10 + ' ROE Result '  +'-'*10)
    print(dfSumROE_All_Sorted)
    # Write to Excel
    sheet_name = 'ALL_ROE_Sorted'
    AddDataToExcelSheet(PathExcelFile, sheet_name, dfSumROE_All_Sorted)
    

'''   
    SumROE_DF_All = pd.DataFrame()
    for RSV_Ndays in range (50, 51, 50):
        for StockHold_Days in range (30, 31, 30):
            # Input:
            StartDate = PickDate('2018_12_01')
            EndDate = PickDate('2018_12_31')
            print('RSV_Ndays: ' + str(RSV_Ndays))
            print('StockHold_Days: ' + str(StockHold_Days))
            
            # Evaluation:
            # equality是累計總報酬率  2018-11-30    0.859861
            # transections是買賣個股的價格 1701     2018-05-03      20.90     ...     2018-05-31       20.55
            equality, transections = backtest(StartDate, EndDate, StockHold_Days, RSV_Ndays, Data())
           
            
            # Output:
            dicData = dic_ROE(equality, RSV_Ndays, StockHold_Days)
            SumROE_DF_All = SumROE_DF_All.append(dicData, ignore_index = True)
    SumROE_DF_All_Sort = SumROE_DF_All.sort_values( by = 'SumROE',  ascending= True)
    print(SumROE_DF_All_Sort)
'''
    
'''   
    # 測試最佳 RSV_Ndays 50 - 200天
    for RSV_Ndays in range(50 , 201,  50):
        for Sold_NDays in range (30, 61,30):
            print('RSV '+str(RSV_Ndays)+' days')
            print('持股天數 '+str(Sold_NDays)+' days')
            data = Data()  
            stocks = mystrategy2(data, RSV_Ndays)
            #equality是累計總報酬率  2018-11-30    0.859861
            #transections是買賣個股的價格 1701     2018-05-03      20.90     ...     2018-05-31       20.55
            equality, transections, SumROE = backtest(datetime.date(2018,1,1), datetime.date(2018,12,1), Sold_NDays, stocks, data)
            #SumROE = round(equality[-1],4)*100-100
            RSV_DF = pd.DataFrame({'RSV_NDays': [RSV_Ndays], '持股天數':[Sold_NDays], '總報酬率':[SumROE]})
    print(RSV_DF)
        
        
         
    # 回測
    #backtest(datetime.date(2018,5,1), datetime.date(2018,12,1), 30, mystrategy2, data)
    
  
    # 利用 mystrategy2 來產生股票清單 stocks
    stocks = mystrategy2(data)
    
    # 用portfolio來幫忙計算，給定 300,000 元，依照今天收盤價，股票張數要如何分配
    # 印出股票資訊
    InvestAmount = 300000
    p, total_invest_money = portfolio(stocks.index, InvestAmount, data)
    
    print('-'*30)
    print('投入 '+ str(InvestAmount) +' \n如何分配股票張數')
    print('-'*30)
    print(p)
    print('total cost: '+str(total_invest_money))    
'''